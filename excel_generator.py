# excel_generator.py - Excel File Creation with UTF-8 Support

import pandas as pd
from utils import get_output_filepath, safe_encode_text, ensure_utf8_environment

def create_excel_file(news_by_source, hours):
    """Save articles to Excel with clickable URL hyperlinks in the specified order and UTF-8 support"""
    # Initialize UTF-8 environment
    ensure_utf8_environment()
    
    if not any(articles for articles in news_by_source.values()):
        print("\nNo articles collected. Exiting...")
        return None, None
    
    # Get file path
    excel_path = get_output_filepath()
    
    # Required columns
    required_columns = ["site", "headline", "link"]
    
    def process_articles_utf8(articles):
        """Process articles to ensure UTF-8 encoding"""
        if not articles:
            return articles
        
        processed_articles = []
        for article in articles:
            # Create a copy to avoid modifying original
            processed_article = article.copy()
            
            # Ensure UTF-8 encoding for all text fields
            text_fields = ['title', 'headline', 'summary', 'description', 'url', 'link', 'site']
            for field in text_fields:
                if field in processed_article and processed_article[field]:
                    processed_article[field] = safe_encode_text(processed_article[field])
            
            # Ensure backward compatibility field mapping
            if 'title' in processed_article and 'headline' not in processed_article:
                processed_article['headline'] = processed_article['title']
            if 'url' in processed_article and 'link' not in processed_article:
                processed_article['link'] = processed_article['url']
                
            processed_articles.append(processed_article)
        
        return processed_articles
    
    try:
        # Import required modules for hyperlinks
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Function to add data to a worksheet with URL hyperlinks
        def add_data_to_sheet(ws, df, sheet_name):
            # Add headers
            ws.append(["Site", "Headline", "Link"])
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for idx, row in df.iterrows():
                site = safe_encode_text(row['site']) if row['site'] else ""
                headline = safe_encode_text(row['headline']) if row['headline'] else ""
                link = safe_encode_text(row['link']) if row['link'] else ""
                
                # Add row data
                ws.append([site, headline, ""])  # Empty link cell initially
                
                # Get the current row number
                row_num = ws.max_row
                
                # Set the link cell with hyperlink formula (single-click)
                link_cell = ws.cell(row=row_num, column=3)
                if link and link.startswith('http'):
                    # Use HYPERLINK formula for reliable single-click behavior
                    link_cell.value = f'=HYPERLINK("{link}", "Click to open")'
                    link_cell.font = Font(color="0000FF", underline="single")
                else:
                    link_cell.value = link
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 25  # Site
            ws.column_dimensions['B'].width = 80  # Headline (wider for readability)
            ws.column_dimensions['C'].width = 15  # Link (just "Click to open")
        
        # 1. Create Singapore sheet first (1st position)
        if 'Singapore' in news_by_source and news_by_source['Singapore']:
            # Process articles for UTF-8 encoding
            processed_articles = process_articles_utf8(news_by_source['Singapore'])
            df_singapore = pd.DataFrame(processed_articles)
            
            # Ensure required columns exist
            for col in required_columns:
                if col not in df_singapore.columns:
                    df_singapore[col] = ""
            df_singapore = df_singapore[required_columns].drop_duplicates(subset=['headline', 'link'])
            
            ws_singapore = wb.create_sheet("Singapore")
            add_data_to_sheet(ws_singapore, df_singapore, "Singapore")
        
        # 2. Create Japan sheet (2nd position)
        if 'Japan' in news_by_source and news_by_source['Japan']:
            # Process articles for UTF-8 encoding
            processed_articles = process_articles_utf8(news_by_source['Japan'])
            df_japan = pd.DataFrame(processed_articles)
            
            # Ensure required columns exist
            for col in required_columns:
                if col not in df_japan.columns:
                    df_japan[col] = ""
            df_japan = df_japan[required_columns].drop_duplicates(subset=['headline', 'link'])
            
            ws_japan = wb.create_sheet("Japan")
            add_data_to_sheet(ws_japan, df_japan, "Japan")
        
        # 3. Create India sheet (3rd position)
        if 'India' in news_by_source and news_by_source['India']:
            # Process articles for UTF-8 encoding
            processed_articles = process_articles_utf8(news_by_source['India'])
            df_india = pd.DataFrame(processed_articles)
            
            # Ensure required columns exist
            for col in required_columns:
                if col not in df_india.columns:
                    df_india[col] = ""
            df_india = df_india[required_columns].drop_duplicates(subset=['headline', 'link'])
            
            ws_india = wb.create_sheet("India")
            add_data_to_sheet(ws_india, df_india, "India")
        
        # 4. Create Korea sheet (4th position)
        if 'Korea' in news_by_source and news_by_source['Korea']:
            # Process articles for UTF-8 encoding
            processed_articles = process_articles_utf8(news_by_source['Korea'])
            df_korea = pd.DataFrame(processed_articles)
            
            # Ensure required columns exist
            for col in required_columns:
                if col not in df_korea.columns:
                    df_korea[col] = ""
            df_korea = df_korea[required_columns].drop_duplicates(subset=['headline', 'link'])
            
            ws_korea = wb.create_sheet("Korea")
            add_data_to_sheet(ws_korea, df_korea, "Korea")
        
        # 5. Create Yahoo Finance sheet (5th position)
        if 'Yahoo Finance' in news_by_source and news_by_source['Yahoo Finance']:
            # Process articles for UTF-8 encoding
            processed_articles = process_articles_utf8(news_by_source['Yahoo Finance'])
            df_yahoo = pd.DataFrame(processed_articles)
            
            # Ensure required columns exist
            for col in required_columns:
                if col not in df_yahoo.columns:
                    df_yahoo[col] = ""
            df_yahoo = df_yahoo[required_columns].drop_duplicates(subset=['headline', 'link'])
            
            ws_yahoo = wb.create_sheet("Yahoo Finance")
            add_data_to_sheet(ws_yahoo, df_yahoo, "Yahoo Finance")
        
        # 6. Add individual sheets for other sources (6th position onwards)
        other_sources = ['TradeWinds', 'Bloomberg', 'TrendForce', 'UDN Money', 'GMK Center']
        
        for source in other_sources:
            if source in news_by_source and news_by_source[source]:
                # Process articles for UTF-8 encoding
                processed_articles = process_articles_utf8(news_by_source[source])
                df_source = pd.DataFrame(processed_articles)
                
                # Ensure required columns exist
                for col in required_columns:
                    if col not in df_source.columns:
                        df_source[col] = ""
                df_source = df_source[required_columns].drop_duplicates(subset=['headline', 'link'])
                
                # Clean sheet name (Excel sheet names have restrictions)
                sheet_name = source.replace('/', '-').replace('\\', '-')
                if len(sheet_name) > 31:  # Excel sheet name limit
                    sheet_name = sheet_name[:31]
                
                ws_source = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws_source, df_source, sheet_name)
        
        # Save the workbook with UTF-8 support
        wb.save(excel_path)
        
        # Calculate totals for summary
        total_articles = sum(len(articles) for articles in news_by_source.values())
        singapore_total = len(news_by_source.get('Singapore', []))
        japan_total = len(news_by_source.get('Japan', [])) 
        india_total = len(news_by_source.get('India', []))
        korea_total = len(news_by_source.get('Korea', []))
        yahoo_total = len(news_by_source.get('Yahoo Finance', []))
        
        # Print summary
        print("-" * 50)
        print("COLLECTION SUMMARY")
        print("-" * 50)
        print(f"Total articles collected: {total_articles}")
        print(f"Collection period: {hours} hours")
        print()
        
        print("Sheet Order & Article Counts:")
        print(f"1. Singapore: {singapore_total} articles")
        print(f"2. Japan: {japan_total} articles")
        print(f"3. India: {india_total} articles") 
        print(f"4. Korea: {korea_total} articles")
        print(f"5. Yahoo Finance: {yahoo_total} articles")
        
        print("\n6th+ position sheets:")
        for source in other_sources:
            if source in news_by_source and news_by_source[source]:
                print(f"   - {source}: {len(news_by_source[source])} articles")
        
        print()
        print(f"✓ Excel saved with UTF-8 support and SINGLE-CLICK links to:")
        print(f"  {excel_path}")
        print(f"  - Sheet 1 'Singapore': {singapore_total} articles")
        print(f"  - Sheet 2 'Japan': {japan_total} articles")
        print(f"  - Sheet 3 'India': {india_total} articles")
        print(f"  - Sheet 4 'Korea': {korea_total} articles (Korean→English translated)")
        print(f"  - Sheet 5 'Yahoo Finance': {yahoo_total} articles")
        for source in other_sources:
            if source in news_by_source and news_by_source[source]:
                sheet_name = source.replace('/', '-').replace('\\', '-')
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                print(f"  - '{sheet_name}' sheet: {len(news_by_source[source])} articles")
        print()
        print("🔗 Link column shows 'Click to open' - single click to open articles!")
        print("🌐 UTF-8 encoding ensures Korean, Chinese, and international characters display correctly!")
        print(f"✓ Output folder:")
        from utils import get_output_directory
        print(f"  {get_output_directory()}")
        
        return True, excel_path
        
    except Exception as e:
        print(f"\nError saving Excel file: {e}")
        print("Make sure you have write permissions in the directory.")
        import traceback
        traceback.print_exc()
        return False, None

if __name__ == "__main__":
    # Test the Excel generator
    sample_data = {
        'Singapore': [
            {'site': 'Test Site', 'headline': 'Test Headline', 'link': 'https://example.com'}
        ],
        'Korea': [
            {'site': 'SEdaily Korea', 'headline': '서울 경제 뉴스', 'link': 'https://example.com'}
        ]
    }
    result = create_excel_file(sample_data, 24)
    print(f"Test result: {result}")